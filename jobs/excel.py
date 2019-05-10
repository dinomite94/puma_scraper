#!/usr/local/bin/python3
from openpyxl import workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import units
from openpyxl.worksheet.cell_range import CellRange

from collections import namedtuple, OrderedDict
import os
import re


HyperLink = namedtuple('HyperLink', ['text', 'link'])
MatchWord = namedtuple('MatchWord', ['column_name', 'match_type', 'match_value'])


root_dir = os.path.dirname(os.path.realpath(__file__))
template_xlsx = os.path.join(root_dir, 'templates/Auswertung Untersuchungskriterien.xlsx')
output_xlsx = os.path.join(root_dir, 'out/jobs/Auswertung Untersuchungskriterien.xlsx')

next_new_id_row = 5

def open():
    xfile = load_workbook(template_xlsx)
    sheet = xfile.get_sheet_by_name('Sheet1')

    column_map = {}
    match_words = set()

    for data in xfile.defined_names.definedName:
        for title, coord in data.destinations:
            if title == 'Sheet1':
                if data.name.startswith('__match_'):
                    cell_range = CellRange(coord)
                    assert cell_range.min_row == cell_range.max_row

                    for column_num in range(cell_range.min_col, cell_range.max_col + 1):
                        (match_value, match_type, *_) = sheet.cell(row=cell_range.min_row, column=column_num).value.lower().split(' ')

                        column_name = '__match_' + match_type + '_' + match_value
                        column_map[column_name] = column_num

                        match_words.add(MatchWord(column_name=column_name, match_type=match_type, match_value=match_value))
                else:
                    column_map[data.name] = sheet[coord][0].column

    return Exporter(xfile=xfile, column_map=column_map, match_words=match_words)



class Exporter:
    def __init__(self, xfile, column_map, match_words):
        self.xfile = xfile
        self.column_map = column_map
        self.match_words = match_words

    def export(self, data_table):
        sheet = self.xfile.get_sheet_by_name('Sheet1')

        def writeCell(cell, data):
            if isinstance(data, list):
                row_num = cell.row
                column_num = cell.column

                for el in data:
                    writeCell(sheet.cell(row=row_num, column=column_num), el)
                    column_num += 1

                return
            elif isinstance(data, tuple) and not isinstance(data, HyperLink):
                if data[1]:
                    cell.comment = Comment(data[1], 'Generator')
                    cell.comment.width = units.points_to_pixels(300)
                data = data[0]

            if isinstance(data, HyperLink):
                cell.value = data.text
                cell.hyperlink = data.link
                cell.style = 'Hyperlink'
            else:
                cell.value = data

            if isinstance(data, int) or isinstance(data, float):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')


        ids = OrderedDict()
        next_new_id_row = 5

        extract_id_re = re.compile(r'r([0-9]+)[._]')
        def extract_id(path):
            match = extract_id_re.findall(path)

            if match:
                return match[0]

        for row in sheet.iter_rows(min_row=next_new_id_row, min_col=2, max_col=2):
            for cell in row:
                if cell.hyperlink:
                    ids[extract_id(cell.hyperlink.target)] = cell.row
                    
                    if cell.row >= next_new_id_row:
                        next_new_id_row = cell.row + 1

        def get_row(id):
            global next_new_id_row
            row = ids.get(id, next_new_id_row)

            if row == next_new_id_row:
                next_new_id_row += 1

            return row

        for row in data_table:
            row_num = get_row(extract_id(row['screenshot_link'].link))

            for key, data in row.items():
                cell = sheet.cell(row=row_num, column=self.column_map[key])

                writeCell(cell, data)

        self.xfile.save(output_xlsx)
